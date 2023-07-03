import time
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import os

def download_songs_from_excel(excel_file):
    # Function to search and download songs
    def search_and_download_song(song_name, artist_name):
        # Initialize the Selenium WebDriver
        driver = webdriver.Firefox()
        driver.maximize_window()

        try:
            # Go to the search page
            driver.get('https://free-mp3-download.net')

            # Find the search input element and enter the search query
            search_input = driver.find_element(By.ID, 'q')
            search_input.send_keys(f'{artist_name} - {song_name}')

            # Find the search button element and click it
            search_button = driver.find_element(By.ID, 'snd')
            search_button.click()

            # Wait for the search results to load
            time.sleep(2)

            # Find the first download button on the search page
            download_button = driver.find_element(By.XPATH, "//button[contains(text(), 'Download')]")
            
            if download_button:
                print("Found the download button!")

                # Click the download button
                download_button.click()

                # Pause execution and allow the user to handle the ad overlay manually
                input("Please handle the ad overlay manually, make sure the captcha is done and that the download is complete. Press Enter to continue...")

                # You can proceed with the download logic here
                # ...

            else:
                print(f"No download option found for: {song_name} - {artist_name}")

        except Exception as e:
            print(f"Error occurred while processing: {song_name} - {artist_name}")
            print(str(e))

        finally:
            # Close the browser
            driver.quit()

    # Load the Excel file
    wb = load_workbook(excel_file)
    ws = wb.active

    # Iterate through the rows in the Excel file
    for row in ws.iter_rows(min_row=2, values_only=True):
        song_name, artist_name = row
        search_and_download_song(song_name, artist_name)

    # Close the Excel file
    wb.close()