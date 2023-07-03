import time
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import os

# Function to search and download songs
def search_and_download_song(song_name, artist_name, output_folder):
    # Initialize the Selenium WebDriver
    driver = webdriver.Firefox()
    driver.maximize_window()

    try:
        # Go to the search page
        driver.get('https://free-mp3-download.net')

        # Find the search input element and enter the search query
        search_input = driver.find_element(By.ID, 'q')
        search_input.send_keys(f'{artist_name} - {song_name}')

        # Find the search button element and click it
        search_button = driver.find_element(By.ID, 'snd')
        search_button.click()

        # Wait for the search results to load
        time.sleep(2)

        # Find the first download button on the search page
        download_button = driver.find_element(By.XPATH, "//button[contains(text(), 'Download')]")
        
        if download_button:
            print("Found the download button!")
            download_button.click()
            # Here a human needs to click on the website. 
            # Because they use a #google_vignette" element which is most likely
            # to prevent automated access.

            
        else:
            print(f"No download option found for: {song_name} - {artist_name}")

    except:
        print(f"No download option found for: {song_name} - {artist_name}")

    finally:
        # Close the browser
        driver.quit()

# Get the directory of the Python script
script_directory = os.path.dirname(os.path.abspath(__file__))

# Load the Excel file
excel_file = os.path.join(script_directory, "playlist_files/playlist_data_2023-06-30_14-02-51.xlsx")
wb = load_workbook(excel_file)
ws = wb.active

# Set the output folder for downloaded songs
output_folder = 'downloaded_songs'

# Iterate through the rows in the Excel file
for row in ws.iter_rows(min_row=2, values_only=True):
    song_name, artist_name = row
    search_and_download_song(song_name, artist_name, output_folder)

# Close the Excel file
wb.close()
